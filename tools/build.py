# -*- coding: utf-8 -*-
"""
build.py — Bộ biên dịch dữ liệu (data compiler)

Quét mọi file trong data/regions/*.json (mỗi file = 1 vùng, là mảng địa điểm)
rồi sinh ra:
  - data/index.json      : manifest (danh sách vùng, thống kê, bbox) cho web app
  - data/bundle.js       : gói dữ liệu cho web app chạy offline (window.RUSSIA_DB)
  - exports/places.csv    : bảng phẳng
  - exports/places.geojson: lớp GeoJSON cho phần mềm GIS
  - exports/places.xlsx   : bảng Excel để lập kế hoạch

TRIẾT LÝ MỞ RỘNG: thêm dữ liệu = thêm/không sửa file trong data/regions/,
rồi chạy lại `python3 tools/build.py`. KHÔNG cần sửa HTML/JS.

Chạy:  python3 tools/build.py
"""
import json, os, csv, glob, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
REGIONS_DIR = os.path.join(ROOT, "data", "regions")
DATA_DIR = os.path.join(ROOT, "data")
EXPORTS_DIR = os.path.join(ROOT, "exports")

REQUIRED = ["id", "name_vi", "region", "coordinates"]


def load_regions():
    regions = []
    files = sorted(glob.glob(os.path.join(REGIONS_DIR, "*.json")))
    for path in files:
        slug = os.path.splitext(os.path.basename(path))[0]
        try:
            with open(path, encoding="utf-8") as f:
                items = json.load(f)
        except Exception as e:
            print(f"  ! Bỏ qua {slug}: lỗi đọc JSON ({e})")
            continue
        if not isinstance(items, list):
            print(f"  ! Bỏ qua {slug}: nội dung không phải mảng")
            continue
        good = []
        for it in items:
            miss = [k for k in REQUIRED if k not in it or it.get(k) in (None, "")]
            if miss:
                print(f"  ! {slug}: bỏ 1 bản ghi thiếu trường {miss} (id={it.get('id')})")
                continue
            good.append(it)
        regions.append({"slug": slug, "items": good})
        print(f"  + {slug}: {len(good)} địa điểm")
    return regions


def bbox_center(items):
    lats = [p["coordinates"]["lat"] for p in items if p.get("coordinates")]
    lons = [p["coordinates"]["lon"] for p in items if p.get("coordinates")]
    if not lats:
        return None, None
    bbox = [min(lons), min(lats), max(lons), max(lats)]  # [w,s,e,n]
    center = [sum(lats) / len(lats), sum(lons) / len(lons)]  # [lat,lon]
    return bbox, center


def build():
    print("Đang biên dịch dữ liệu...")
    regions = [r for r in load_regions() if r["items"]]  # bỏ vùng rỗng (vd tỉnh cũ đã gộp đi)
    all_places = []
    region_meta = []
    cat_counts = {}

    for r in regions:
        items = r["items"]
        for p in items:
            all_places.append(p)
            for c in p.get("categories", []):
                cat_counts[c] = cat_counts.get(c, 0) + 1
        bbox, center = bbox_center(items)
        name_vi = items[0].get("region_name_vi", r["slug"]) if items else r["slug"]
        fed = items[0].get("federal_district") if items else None
        country = (items[0].get("country") if items else None) or "russia"
        region_meta.append({
            "slug": r["slug"],
            "name_vi": name_vi,
            "federal_district": fed,
            "country": country,
            "count": len(items),
            "bbox": bbox,
            "center": center,
        })

    # bbox toàn cục
    gbbox, gcenter = bbox_center(all_places)
    generated = "BUILD_TIME"  # thay bằng thời gian thực khi chạy (xem cuối hàm)
    try:
        generated = datetime.datetime.now().isoformat(timespec="seconds")
    except Exception:
        pass

    country_counts = {}
    country_names = {"russia": "Nga", "vietnam": "Việt Nam"}
    for p in all_places:
        c = p.get("country") or "russia"
        country_counts[c] = country_counts.get(c, 0) + 1

    index = {
        "generated": generated,
        "total_places": len(all_places),
        "regions": region_meta,
        "categories": cat_counts,
        "countries": country_counts,
        "country_names": country_names,
        "bbox": gbbox,
        "center": gcenter,
    }

    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    # 1) index.json
    with open(os.path.join(DATA_DIR, "index.json"), "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    # 2) bundle.js  (cho web app chạy bằng file:// — không cần server)
    bundle = {"meta": index, "places": all_places}
    with open(os.path.join(DATA_DIR, "bundle.js"), "w", encoding="utf-8") as f:
        f.write("/* Tự động sinh bởi build.py — KHÔNG sửa tay. */\n")
        f.write("window.RUSSIA_DB = ")
        json.dump(bundle, f, ensure_ascii=False)
        f.write(";\n")

    # 3) exports/places.json (tiện cho lập trình / API sau này)
    with open(os.path.join(EXPORTS_DIR, "places.json"), "w", encoding="utf-8") as f:
        json.dump(all_places, f, ensure_ascii=False, indent=2)

    # 4) exports/places.csv
    csv_cols = [
        "id", "name_vi", "name_ru", "name_en", "region_name_vi", "categories",
        "lat", "lon", "rating_value", "rating_count", "rating_source",
        "duration", "ticket", "hours", "best_time",
        "yandex_map", "google_map", "official_site", "status", "short_presentation",
    ]
    with open(os.path.join(EXPORTS_DIR, "places.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(csv_cols)
        for p in all_places:
            pr = p.get("practical", {}) or {}
            rt = p.get("rating", {}) or {}
            w.writerow([
                p.get("id"), p.get("name_vi"), p.get("name_ru"), p.get("name_en"),
                p.get("region_name_vi"), "|".join(p.get("categories", [])),
                p["coordinates"]["lat"], p["coordinates"]["lon"],
                rt.get("value"), rt.get("count"), rt.get("source"),
                pr.get("duration_vi"), pr.get("ticket_vi"), pr.get("hours_vi"), pr.get("best_time_vi"),
                (p.get("maps", {}) or {}).get("yandex"), (p.get("maps", {}) or {}).get("google"),
                p.get("official_site"), p.get("status"), p.get("presentation_short_vi"),
            ])

    # 5) exports/places.geojson
    features = []
    for p in all_places:
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [p["coordinates"]["lon"], p["coordinates"]["lat"]]},
            "properties": {
                "id": p.get("id"), "name_vi": p.get("name_vi"), "name_ru": p.get("name_ru"),
                "name_en": p.get("name_en"), "region": p.get("region_name_vi"),
                "categories": p.get("categories", []),
                "rating": (p.get("rating", {}) or {}).get("value"),
                "reviews": (p.get("rating", {}) or {}).get("count"),
                "short": p.get("presentation_short_vi"),
                "yandex": (p.get("maps", {}) or {}).get("yandex"),
                "google": (p.get("maps", {}) or {}).get("google"),
                "photo": p.get("photo"), "status": p.get("status"),
            },
        })
    fc = {"type": "FeatureCollection", "features": features}
    with open(os.path.join(EXPORTS_DIR, "places.geojson"), "w", encoding="utf-8") as f:
        json.dump(fc, f, ensure_ascii=False, indent=2)

    # 6) exports/places.xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Places"
        headers = [
            "ID", "Tên (VI)", "Tên (RU)", "Tên (EN)", "Vùng", "Loại",
            "Vĩ độ", "Kinh độ", "Sao", "Số đánh giá", "Nguồn ĐG",
            "Thời lượng", "Giá vé", "Giờ mở cửa", "Thời điểm đẹp",
            "Yandex Map", "Google Map", "Web chính thức", "Trạng thái", "Thuyết trình ngắn",
        ]
        ws.append(headers)
        hfill = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(color="FFFFFF", bold=True)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for p in all_places:
            pr = p.get("practical", {}) or {}
            rt = p.get("rating", {}) or {}
            ws.append([
                p.get("id"), p.get("name_vi"), p.get("name_ru"), p.get("name_en"),
                p.get("region_name_vi"), ", ".join(p.get("categories", [])),
                p["coordinates"]["lat"], p["coordinates"]["lon"],
                rt.get("value"), rt.get("count"), rt.get("source"),
                pr.get("duration_vi"), pr.get("ticket_vi"), pr.get("hours_vi"), pr.get("best_time_vi"),
                (p.get("maps", {}) or {}).get("yandex"), (p.get("maps", {}) or {}).get("google"),
                p.get("official_site"), p.get("status"), p.get("presentation_short_vi"),
            ])
        widths = [22, 30, 24, 24, 18, 16, 10, 10, 6, 12, 12, 14, 22, 26, 22, 30, 30, 26, 12, 50]
        for i, wdt in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = wdt
        ws.freeze_panes = "A2"
        wb.save(os.path.join(EXPORTS_DIR, "places.xlsx"))
    except Exception as e:
        print(f"  ! Không tạo được XLSX: {e}")

    print(f"Xong. Tổng {len(all_places)} địa điểm, {len(region_meta)} vùng.")
    print("Đã sinh: data/index.json, data/bundle.js, exports/{places.json,csv,geojson,xlsx}")
    return index


if __name__ == "__main__":
    build()
